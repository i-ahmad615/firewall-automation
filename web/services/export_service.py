"""Export alerts / firewall actions / logs to CSV or Excel (.xlsx)."""
from __future__ import annotations

import csv
import io
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core import database

_TABLE_HEADERS: dict[str, list[str]] = {
    "alerts": [
        "id", "received_at", "subject", "sender", "origin_ip", "classification",
        "status", "action_taken", "reason", "notification_sent",
    ],
    "firewall_actions": [
        "id", "occurred_at", "ip", "rule_name", "result", "duplicate",
        "allowed_list", "notification_sent", "status", "detail",
    ],
    "logs": ["id", "timestamp", "severity", "module", "action", "message"],
}

ExportTable = Literal["alerts", "firewall_actions", "logs"]


def _rows_for(table: ExportTable) -> list[dict[str, Any]]:
    return database.all_rows(table)


def to_csv(table: ExportTable) -> bytes:
    headers = _TABLE_HEADERS[table]
    rows = _rows_for(table)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx(table: ExportTable) -> bytes:
    headers = _TABLE_HEADERS[table]
    rows = _rows_for(table)

    wb = Workbook()
    ws = wb.active
    ws.title = table[:31]

    header_fill = PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header))

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(header)] + [len(str(r.get(header, ""))) for r in rows]) if rows else len(header)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
